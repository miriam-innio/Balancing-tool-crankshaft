import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter



#-----------------------------------------------------------------------------------------------------------------------
#                                       0. Auxiliary functions for unit conversion
#-----------------------------------------------------------------------------------------------------------------------


def units_speed(units):
    return 1 if units == "rpm" else None

def units_number(units):
    return 1 if units == "#" else None

def units_none(units):
    return 1 if units == "-" else None

def units_angle(units):
    fact = None
    if units == "deg":
        fact = np.pi / 180
    elif units == "rad":
        fact = 1
    return fact

def units_mass(units):
    fact = None
    if units == "lb":
        fact = 0.45359237
    elif units == "kg":
        fact = 1
    return fact

def units_length(units):
    fact = None
    if units == "in":
        fact = 0.0254
    elif units == "cm":
        fact = 1e-2
    elif units == "mm":
        fact = 1e-3
    elif units == "m":
        fact = 1
    return fact

def read_scalar(data, indrow, variable, name):
    units = data.iloc[indrow, 1]
    if variable == "speed":
        fact = units_speed(units)
    elif variable == "number":
        fact = units_number(units)
    elif variable == "angle":
        fact = units_angle(units)
    elif variable == "none":
        fact = units_none(units)
    elif variable == "mass":
        fact = units_mass(units)
    elif variable == "length":
        fact = units_length(units)
    else:
        raise ValueError(f"Variable '{variable}' is not supported.")

    if fact is None:
        raise ValueError(f"Units '{units}' for variable '{name}' are not supported.")

    return fact * data.iloc[indrow, 2]

#-----------------------------------------------------------------------------------------------------------------------
#                                           1. Read data from Excel
#-----------------------------------------------------------------------------------------------------------------------

def read_text(data, indrow, expected_unit="-", name=""):
    units = data.iloc[indrow, 1]
    if units != expected_unit:
        raise ValueError(f"Units '{units}' for variable '{name}' are not supported.")
    return data.iloc[indrow, 2]


def read_input_excel(file_path, sheet_name="Balancing"):
    # Table 1
    data = pd.read_excel(file_path, sheet_name=sheet_name, usecols="B:D", skiprows=3, nrows=4, header=None)

    speed = read_scalar(data, 0, "speed", "Speed")
    n_cyl = int(read_scalar(data, 1, "number", "Number of cylinders"))
    cyl_arr = read_text(data, 2, expected_unit="-", name="Cylinder arrangement")
    vee_rad = read_scalar(data, 3, "angle", "Vee angle")
    vee_deg = vee_rad * 180 / np.pi

    eng_conf = {
        "speed": speed,
        "n_cyl": n_cyl,
        "cyl_arr": cyl_arr,
        "vee": vee_deg,
    }

    # Table 2
    data = pd.read_excel(file_path, sheet_name=sheet_name, usecols="F:H", skiprows=3, nrows=7, header=None)

    crankshaft_geo = {
        "stroke": read_scalar(data, 0, "length", "Stroke"),
        "l_conrod": read_scalar(data, 1, "length", "Conrod length"),
        "r_web": read_scalar(data, 2, "length", "Web CoG"),
        "r_cw": read_scalar(data, 3, "length", "Counterweight CoG"),
        "dist_main_main": read_scalar(data, 4, "length", "Distance between main journals"),
        "dist_main_cyl1": read_scalar(data, 5, "length", "Distance from main journal to cylinder 1"),
        "dist_main_cyl2": read_scalar(data, 6, "length", "Distance from main journal to cylinder 2"),
    }

    # Table 3
    data = pd.read_excel(file_path, sheet_name=sheet_name, usecols="J:L", skiprows=3, nrows=6, header=None)

    masses = {
        "mass_piston": read_scalar(data, 0, "mass", "Piston assembly mass"),
        "mass_con_se": read_scalar(data, 1, "mass", "Conrod oscillating mass (small end)"),
        "mass_con_be": read_scalar(data, 2, "mass", "Conrod rotating mass (big end)"),
        "mass_pin": read_scalar(data, 3, "mass", "Crank pin mass"),
        "mass_web": read_scalar(data, 4, "mass", "Web mass"),
        "mass_cw": read_scalar(data, 5, "mass", "Counterweight mass"),
    }

    # Table 4
    n_throw = n_cyl // 2 if cyl_arr == "Vee" else n_cyl
    throw_info = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        usecols=range(13, 13 + n_throw + 1),
        skiprows=3,
        header=None,
    ).dropna(how="all")

    case_number = throw_info.iloc[:, 0].to_numpy()
    throw_angles_mat = throw_info.iloc[:, 1:].to_numpy()

    if len(case_number) != len(set(case_number)):
        raise ValueError("Case number column has repeated ids.")

    throw_angles_mat = throw_angles_mat.T

    return {
        "eng_conf": eng_conf,
        "crankshaft_geo": crankshaft_geo,
        "masses": masses,
        "case_number": case_number,
        "throw_angles_mat": throw_angles_mat,
    }


def build_results_dataframe(case_number, ratio_r, fry, frz, mry, mrz, foy, foz, moy, moz):
    fr = np.sqrt(fry ** 2 + frz ** 2)
    mr = np.sqrt(mry ** 2 + mrz ** 2)

    data = {
        "Case number": case_number,
        "Balancing ratio": ratio_r,
        "F_r (N)": np.max(np.abs(fr), axis=0),
        "M_r (Nmm)": np.max(np.abs(mr), axis=0) * 1e3,
        "F_o 1H (N)": np.max(np.abs(foy[0, :, :]), axis=0),
        "F_o 1V (N)": np.max(np.abs(foz[0, :, :]), axis=0),
        "M_o 1H (Nmm)": np.max(np.abs(moy[0, :, :]), axis=0) * 1e3,
        "M_o 1V (Nmm)": np.max(np.abs(moz[0, :, :]), axis=0) * 1e3,
        "F_o 2H (N)": np.max(np.abs(foy[1, :, :]), axis=0),
        "F_o 2V (N)": np.max(np.abs(foz[1, :, :]), axis=0),
        "M_o 2H (Nmm)": np.max(np.abs(moy[1, :, :]), axis=0) * 1e3,
        "M_o 2V (Nmm)": np.max(np.abs(moz[1, :, :]), axis=0) * 1e3,
        "F_o 4H (N)": np.max(np.abs(foy[3, :, :]), axis=0),
        "F_o 4V (N)": np.max(np.abs(foz[3, :, :]), axis=0),
        "M_o 4H (Nmm)": np.max(np.abs(moy[3, :, :]), axis=0) * 1e3,
        "M_o 4V (Nmm)": np.max(np.abs(moz[3, :, :]), axis=0) * 1e3,
        "F_o 6H (N)": np.max(np.abs(foy[5, :, :]), axis=0),
        "F_o 6V (N)": np.max(np.abs(foz[5, :, :]), axis=0),
        "M_o 6H (Nmm)": np.max(np.abs(moy[5, :, :]), axis=0) * 1e3,
        "M_o 6V (Nmm)": np.max(np.abs(moz[5, :, :]), axis=0) * 1e3,
        "F_o 8H (N)": np.max(np.abs(foy[7, :, :]), axis=0),
        "F_o 8V (N)": np.max(np.abs(foz[7, :, :]), axis=0),
        "M_o 8H (Nmm)": np.max(np.abs(moy[7, :, :]), axis=0) * 1e3,
        "M_o 8V (Nmm)": np.max(np.abs(moz[7, :, :]), axis=0) * 1e3,
    }

    return pd.DataFrame(data)


def write_results_excel(results_df, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header
    for col_idx, col_name in enumerate(results_df.columns, start=2):
        ws.cell(row=2, column=col_idx, value=col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    # Data
    for row_idx, row in enumerate(results_df.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, np.integer)):
                cell.number_format = "0"
            elif isinstance(value, (float, np.floating)):
                cell.number_format = "0.0E+0"

    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = centered

    white_side = Side(style="thin", color="FFFFFF")
    black_side = Side(style="thin", color="000000")
    white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)

    last_row = 2 + len(results_df)
    last_col = 1 + len(results_df.columns)

    for r in range(2, last_row + 1):
        for c in range(2, last_col + 1):
            ws.cell(r, c).border = white_border

    def set_outer_border(ws, row1, col1, row2, col2, side):
        for r in range(row1, row2 + 1):
            for c in range(col1, col2 + 1):
                cell = ws.cell(r, c)
                left = cell.border.left
                right = cell.border.right
                top = cell.border.top
                bottom = cell.border.bottom
                if c == col1:
                    left = side
                if c == col2:
                    right = side
                if r == row1:
                    top = side
                if r == row2:
                    bottom = side
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    set_outer_border(ws, 2, 2, last_row, last_col, black_side)
    set_outer_border(ws, 2, 2, 2, last_col, black_side)

    wb.save(output_path)